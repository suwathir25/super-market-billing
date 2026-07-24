from flask import Blueprint, render_template, request, make_response
from app.views import login_required, role_required
from app.models.database import query_db
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sales_bp = Blueprint('sales', __name__, url_prefix='/sales')


@sales_bp.route('/')
@login_required
@role_required('admin', 'manager')
def index():
    search = request.args.get('search', '').strip()
    payment_filter = request.args.get('payment', '').strip()
    status_filter = request.args.get('status', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    per_page = 15

    base_query = """
        SELECT b.id, b.bill_number, b.total_amount, b.subtotal, b.discount_amount,
               b.tax_amount, b.payment_method, b.status, b.created_at,
               COALESCE(c.name, 'Walk-in Customer') as customer_name,
               u.username as cashier
        FROM bills b
        LEFT JOIN customers c ON b.customer_id = c.id
        LEFT JOIN users u ON b.user_id = u.id
        WHERE 1=1
    """
    params = []

    if search:
        base_query += " AND (b.bill_number LIKE ? OR c.name LIKE ?)"
        params += [f'%{search}%', f'%{search}%']
    if payment_filter:
        base_query += " AND b.payment_method = ?"
        params.append(payment_filter)
    if status_filter:
        base_query += " AND b.status = ?"
        params.append(status_filter)

    total_count = query_db(
        f"SELECT COUNT(*) as cnt FROM ({base_query}) t", params, one=True
    )
    total_bills = total_count['cnt'] if total_count else 0
    total_pages = max(1, (total_bills + per_page - 1) // per_page)

    base_query += " ORDER BY b.created_at DESC LIMIT ? OFFSET ?"
    params += [per_page, (page - 1) * per_page]
    bills = query_db(base_query, params)

    # Summary stats
    summary = query_db("""
        SELECT
            COUNT(*) as total_count,
            COALESCE(SUM(total_amount), 0) as total_revenue,
            COALESCE(SUM(CASE WHEN status = 'completed' THEN total_amount ELSE 0 END), 0) as completed_revenue
        FROM bills
    """, one=True)

    currency_row = query_db("SELECT value FROM settings WHERE key='currency'", one=True)
    currency = currency_row['value'] if currency_row else '₹'

    return render_template('sales.html',
                           bills=bills, summary=summary, currency=currency,
                           search=search, payment_filter=payment_filter,
                           status_filter=status_filter,
                           page=page, total_pages=total_pages)


@sales_bp.route('/export')
@login_required
@role_required('admin', 'manager')
def export():
    search = request.args.get('search', '').strip()
    payment_filter = request.args.get('payment', '').strip()
    status_filter = request.args.get('status', '').strip()

    base_query = """
        SELECT b.id, b.bill_number, b.total_amount, b.subtotal, b.discount_amount,
               b.tax_amount, b.payment_method, b.status, b.created_at,
               b.cash_amount, b.card_amount, b.net_banking_amount,
               COALESCE(c.name, 'Walk-in Customer') as customer_name,
               u.username as cashier
        FROM bills b
        LEFT JOIN customers c ON b.customer_id = c.id
        LEFT JOIN users u ON b.user_id = u.id
        WHERE 1=1
    """
    params = []

    if search:
        base_query += " AND (b.bill_number LIKE ? OR c.name LIKE ?)"
        params += [f'%{search}%', f'%{search}%']
    if payment_filter:
        base_query += " AND b.payment_method = ?"
        params.append(payment_filter)
    if status_filter:
        base_query += " AND b.status = ?"
        params.append(status_filter)

    base_query += " ORDER BY b.created_at DESC"
    bills = query_db(base_query, params)

    currency_row = query_db("SELECT value FROM settings WHERE key='currency'", one=True)
    currency_sym = currency_row['value'] if currency_row else '₹'
    money_fmt = '#,##0.00'

    # ── Workbook setup ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payments Export'

    # Style helpers
    header_fill   = PatternFill('solid', fgColor='1E293B')   # dark slate
    alt_fill      = PatternFill('solid', fgColor='F1F5F9')   # light blue-grey
    total_fill    = PatternFill('solid', fgColor='0F172A')   # deeper slate
    thin_border   = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    body_font     = Font(name='Calibri', size=10)
    total_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    right_align   = Alignment(horizontal='right',  vertical='center')
    left_align    = Alignment(horizontal='left',   vertical='center')

    # ── Title row ───────────────────────────────────────────────────────────────
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = 'SuperMart — Payments Transaction Report'
    title_cell.font  = Font(name='Calibri', bold=True, size=14, color='1E293B')
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # ── Column headers (row 2) ─────────────────────────────────────────────────
    COLUMNS = [
        ('Bill Number',    18, 'left'),
        ('Date & Time',    20, 'center'),
        ('Customer Name',  24, 'left'),
        ('Cashier',        14, 'left'),
        ('Payment Method', 16, 'center'),
        ('Status',         12, 'center'),
        ('Subtotal',       14, 'right'),
        ('Discount',       13, 'right'),
        ('Tax',            12, 'right'),
        ('Total Amount',   15, 'right'),
        ('Cash Paid',      13, 'right'),
        ('Card Paid',      13, 'right'),
        ('UPI Paid',       12, 'right'),
    ]

    for col_idx, (title, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'   # freeze title + header

    # ── Data rows ──────────────────────────────────────────────────────────────
    MONEY_COLS = {7, 8, 9, 10, 11, 12, 13}   # 1-based column indices

    for row_idx, b in enumerate(bills, start=3):
        row_fill = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        row_data = [
            b['bill_number'],
            b['created_at'][:16] if b['created_at'] else '',
            b['customer_name'],
            b['cashier'] or '',
            b['payment_method'],
            b['status'].capitalize(),
            b['subtotal'],
            b['discount_amount'],
            b['tax_amount'],
            b['total_amount'],
            b['cash_amount'] or 0.0,
            b['card_amount'] or 0.0,
            b['net_banking_amount'] or 0.0,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = body_font
            cell.fill   = row_fill
            cell.border = thin_border
            _, _, align = COLUMNS[col_idx - 1]
            if align == 'right':
                cell.alignment = right_align
            elif align == 'center':
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if col_idx in MONEY_COLS:
                cell.number_format = money_fmt

        ws.row_dimensions[row_idx].height = 18

    # ── Totals summary row ─────────────────────────────────────────────────────
    total_row = ws.max_row + 1
    ws.row_dimensions[total_row].height = 22

    ws.cell(row=total_row, column=1, value='TOTALS').font = total_font
    ws.cell(row=total_row, column=1).fill      = total_fill
    ws.cell(row=total_row, column=1).alignment = center_align
    ws.cell(row=total_row, column=1).border    = thin_border

    for col_idx in range(2, 14):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill   = total_fill
        cell.border = thin_border
        if col_idx in MONEY_COLS:
            # SUM formula spanning data rows
            data_start = 3
            data_end   = total_row - 1
            col_letter = get_column_letter(col_idx)
            cell.value         = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
            cell.font          = total_font
            cell.number_format = money_fmt
            cell.alignment     = right_align

    # ── Stream to response ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=payments_export.xlsx'
    return response

